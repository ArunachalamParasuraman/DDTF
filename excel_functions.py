from openpyxl import load_workbook

class Arun_excel_function:

    def __init__(self,excel_file_name,sheet_name):
        self.file = excel_file_name
        self.sheet = sheet_name
    
    def Row_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return(sheet.max_row)
    
    def Column_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return(sheet.max_column)
    
    def Read_Data(self,row_number,column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return(sheet.cell(row = row_number,column = column_number).value)
    
    def Write_Data(self,row_number,column_number,data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        data = sheet.cell(row = row_number,column = column_number).value
        workbook.save(self.file)